import customtkinter as tk
from tkinter import *
from tkinter import messagebox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib
import sqlite3

from PIL import Image, ImageTk

#============================================= aparecia do app ============================================================= 
tk.set_appearance_mode('system')
tk.set_default_color_theme('blue')

#============================================= TELA DE INICI=================================================================

janela = tk.CTk()# cria a janela principal em seguida cria mainloop juntos.
janela. title('APP0.1 victor')#colocar titulo no 
janela.geometry("800x550+350+50 ")# tamanha da janela principal.
janela.maxsize(width=900,height=550)# Só pode espander para esse tamanho colocado.
janela.minsize(width=500,height=250)# Só pode diminuir para esse tamanho colocado.


#=================================================== TELA PRINCIPAL =================================================================

def tela_p():
    janela.destroy() #apagando a tela de login apos o aparecimento da tela principal 
    janela_p = tk.CTk()
    janela_p.title('APP 0.1')
    janela_p.geometry('800x550+350+50')
    janela_p.maxsize(width=900,height=550) 
    janela_p.minsize(width=500,height=250)
   
#----------------------------- opçoes da tabela principal--------------------------------------------------------.
    opcao=tk.CTkTabview(janela_p,width=450, height=500,border_width=2,
                                   corner_radius=10,border_color='light green',segmented_button_unselected_color='green',
                                   segmented_button_fg_color='light green',segmented_button_unselected_hover_color='light green',
                                   segmented_button_selected_color='dark green',segmented_button_selected_hover_color='light green')
    opcao.add('OPÇÃO 1') 
    opcao.add('OPÇÃO 2')
    opcao.add('OPÇÃO 3')
    opcao.add('RESULT')
    opcao.tab('OPÇÃO 1')
    opcao.tab('OPÇÃO 2')
    opcao.tab('OPÇÃO 3')
    opcao.tab('RESULT')
    opcao.pack()
#=================================================== OPÇÃO 1 =================================================================

    opcao1 = tk.CTkLabel(opcao.tab('OPÇÃO 1'),text="CALCULO METABOLICO\n"'\n'"São fórmulas matemáticas que levam em consideração\n"
                             "fatores como idade, sexo e nível de atividade física\n"
                             "para estimar o gasto calórico isso ajuda a criar.\n"
                             "estrategias para subir ou baixar seu peso \n",font=('system',8))
    opcao1.place(x=41,y=15)
    
    sexo_var = tk.IntVar()
    peso_var = tk.IntVar()
    altura_var = tk.IntVar()
    idade_var = tk.IntVar()
#--------------------------função metabolismo-------------------------------------------------
  
    def calcular_metabolismo():
        
      try:
        sexo_mt = sexo_var.get()
        peso = peso_var.get()
        altura = altura_var.get()
        idade = idade_var.get()
        

        if sexo_mt == 1:
           labol = tk.CTkLabel(opcao.tab('OPÇÃO 1'))
           labol.place(x=75,y=360)
           resultado1 = (f'{88.362 + (13.397 *float(peso)) + (4.799 * float(altura)* 100) - (5.677 * float(idade)):.3f}')
          
           exibir = tk.CTkLabel(opcao.tab('OPÇÃO 1'),text=f'Seu organismo gasta para sobreviver: {resultado1} [KCAL]',font=('system',12))
           exibir.place(x=40,y=360)

           meta_resul= tk.CTkLabel(opcao.tab('RESULT'),text=resultado1,font=('impact',15))#EXIBIR RESULTADO NA TELA DE RESULTADO 
           meta_resul.place(x=110,y=60)
         
           
        elif sexo_mt == 2:
         resultado = (f'{447.593 + (9.247 * float(peso)) + (3.098 * float(altura) * 100) - (4.330 * float(idade)):.3f} ')
         
         exibir = tk.CTkLabel(opcao.tab('OPÇÃO 1'),text=f'Seu organismo gasta para sobreviver: {resultado} [KCAL]',font=('system',12))
         exibir.place(x=40,y=360)
 
         meta_resulr = tk.CTkLabel(opcao.tab('RESULT'),text=resultado,font=('impact',15))
         meta_resulr.place(x=110,y=60)
         

        
      except Exception: #pega qualquer tipo de erro

          resu_META = (f'(Preencha corretamente os campos acima  !!!!!!)')
          print(resu_META)

          erro_META = tk.CTkLabel(opcao.tab('OPÇÃO 1'),text='ERRO  : ',text_color='RED',font=('system',12.2))
          erro_META.place(x=40,y=360)

          exibir_META = tk.CTkLabel(opcao.tab('OPÇÃO 1'),text=resu_META,font=('system',12))
          exibir_META.place(x=95,y=360)

#==================================================================================================

    idade_var = tk.CTkEntry(opcao.tab('OPÇÃO 1'), fg_color='white', placeholder_text='Idade...', font=('System', 15),
                                   width=250, border_color='light green', border_width=2, corner_radius=15,
                                   height=35, text_color='black')
    idade_var.place(x=90, y=150)

    peso_var = tk.CTkEntry(opcao.tab('OPÇÃO 1'), fg_color='white', placeholder_text='Peso...', font=('System', 15),
                                  width=250, border_color='light green', border_width=2, corner_radius=15,
                                  height=35, text_color='black')
    peso_var.place(x=90, y=190)

    altura_var = tk.CTkEntry(opcao.tab('OPÇÃO 1'), fg_color='white', placeholder_text=' Altura...',
                                     font=('System', 15),
                                     width=250, border_color='light green', border_width=3, corner_radius=15,
                                     height=35, text_color='black')
    altura_var.place(x=90, y=230)

    sexo_M_mt = tk.CTkRadioButton(opcao.tab('OPÇÃO 1'), text='MASCULINO',fg_color='dark green',hover_color='light green', variable=sexo_var, value=1)
    sexo_F_mt = tk.CTkRadioButton(opcao.tab('OPÇÃO 1'), text='FEMININO',fg_color='dark green',hover_color='light green',variable=sexo_var, value=2)
    sexo_M_mt.place(x=110, y=280)
    sexo_F_mt.place(x=225, y=280)

    botao_op1 = tk.CTkButton(opcao.tab('OPÇÃO 1'), text='>>START<<', fg_color='dark green',hover_color='light green',width=160,height=35,
                                        border_width=2, corner_radius=9,font=('system',12) ,command=calcular_metabolismo)
    botao_op1.place(x=130,y=320)
#================================================= fim da opção 1 ============================================================

#=================================================== OPÇÃO 2 =================================================================
    texto=tk.CTkLabel(opcao.tab('OPÇÃO 2'),text='ÍNDICE DE MASSA CORPORIA',font=('system',12))
    texto.place(x=115,y=10)

    opcao2 = tk.CTkLabel(opcao.tab('OPÇÃO 2'),text='O índice de massa corporal (IMC)\n' 
          
                               'é uma medida internacional usada para calcular\nse uma pessoa está no peso ideal.',font=('system',12))
    opcao2.place(x=60,y=60)

    #--------------------------------------- FUNÇÃO IMC----------------------------------------------------
    sexo_imc = tk.IntVar(value=0)
    peso_imc = tk.IntVar()
    altura_imc = tk.IntVar()

    def calculo_imc():
      try:
       peso_imc_var = peso_imc.get()
       altura_imc_var = altura_imc.get()
       sexo_imc_var = sexo_imc.get()
      
       if sexo_imc_var == 1:
         resu_imc = (f'{float(peso_imc_var)/(float(altura_imc_var)):.2f} ') 
         if(float(peso_imc_var)/(float(altura_imc_var)*2))< 19.1 :
            tipo_IMC = ('  "  ABAIXO DO PESO  " ')
         elif(float(peso_imc_var)/(float(altura_imc_var)*2))>= 19.1 and (float(peso_imc_var)/(float(altura_imc_var)*2))<= 25.8:
            tipo_IMC = ('  "    PESO NORMAL  "  ')
         elif (float(peso_imc_var)/(float(altura_imc_var)*2))>= 25.8 and (float(peso_imc_var)/(float(altura_imc_var)*2))<=27.3:
            tipo_IMC = ('  "  ACIMA DO PESO  " ')   
         elif (float(peso_imc_var)/(float(altura_imc_var)*2))>=32.4:
            tipo_IMC = ('"RISCO DE OBSIDADE"')
    
         exibir_imc = tk.CTkLabel(opcao.tab('OPÇÃO 2'),text=f' O SEU (IMC) É DE {resu_imc}',font=('system',12.5))
         exibir_imc.place(x=15,y=330)
         imc_result = tk.CTkLabel(opcao.tab('RESULT'),text=f'{resu_imc}',font=('impact',15))#EXIBIR RESULTADO NA TELA DE RESULTADO 
         imc_result.place(x=50,y=260)

         exibir_tipo = tk.CTkLabel(opcao.tab('OPÇÃO 2'),text=tipo_IMC,font=('system',12.5))
         exibir_tipo.place(x=165,y=330)
         
       elif sexo_imc_var == 2:
          resu_imc=(f'{float(peso_imc_var)/(float(altura_imc_var)*2):.2f}  ')
          if(float(peso_imc_var)/(float(altura_imc_var)*2))< 18.5 :
            tipo_IMC = ('  "  ABAIXO DO PESO  "  ')
          elif(float(peso_imc_var)/(float(altura_imc_var)*2))>= 18.5 and (float(peso_imc_var)/(float(altura_imc_var)*2))<=24.9:
            tipo_IMC = ('   "   PESO NORMAL   "  ')
          elif(float(peso_imc_var)/(float(altura_imc_var)*2))>= 24.9 and (float(peso_imc_var)/(float(altura_imc_var)*2))<=29.9:
            tipo_IMC = ('  "  ACIMA DO PESO  "  ')   
          elif (float(peso_imc_var)/(float(altura_imc_var)*2))>=30:
            tipo_IMC = ('"RISCO DE OBSIDADE"')

          exibir_imc = tk.CTkLabel(opcao.tab('OPÇÃO 2'),text=f' O SEU (IMC) É DE {resu_imc}',font=('system',12.5))
          exibir_imc.place(x=15,y=330)
          imc_result = tk.CTkLabel(opcao.tab('RESULT'),text=f'{resu_imc}',font=('impact',15))#EXIBIR RESULTADO NA TELA DE RESULTADO 
          imc_result.place(x=50,y=260)

          exibir_tipo = tk.CTkLabel(opcao.tab('OPÇÃO 2'),text=tipo_IMC,font=('system',12.5))
          exibir_tipo.place(x=165,y=330)



      except Exception:
          resu_imc = (f' Preencha os campos corretamente !!!  ')
          exibir_imc = tk.CTkLabel(opcao.tab('OPÇÃO 2'),text=f'{resu_imc}',font=('system',12))
          exibir_imc.place(x=65,y=330)
          
          erro_imc = (' ERRO :')
          erro_imc = tk.CTkLabel(opcao.tab('OPÇÃO 2'),text=erro_imc,text_color='RED',font=('system',12))
          erro_imc.place(x=15,y=330)
#========================================================================================================================
    peso_imc = tk.CTkEntry(opcao.tab('OPÇÃO 2'),fg_color='white', placeholder_text='Peso',font=('System',15),
                              width=250,border_color='light green', border_width=2,corner_radius=15,
                              height=35, text_color='black')
    peso_imc.place(x=90,y=140)

    altura_imc = tk.CTkEntry(opcao.tab('OPÇÃO 2'),fg_color='white', placeholder_text='Altura',font=('System',15),
                              width=250,border_color='light green', border_width=2,corner_radius=15,
                              height=35, text_color='black')
    altura_imc.place(x=90,y=180)

    sexo_imc_M = tk.CTkRadioButton(opcao.tab('OPÇÃO 2'),text='FEMININO',hover_color='light green',fg_color='dark green',variable=sexo_imc,value=1)
    sexo_imc_M.place(x=225,y=230)
    sexo_imc_F = tk.CTkRadioButton(opcao.tab('OPÇÃO 2'),text='MASCULINO',hover_color='light green',fg_color='dark green',variable=sexo_imc,value=2)
    sexo_imc_F.place(x=115,y=230)

     
    botao_imc = tk.CTkButton(opcao.tab('OPÇÃO 2'),text='>>START<<',hover_color='light green',fg_color='dark green',font=('System',15),border_width=2,
                           corner_radius=9,width=160,height=35,command=calculo_imc)
    botao_imc.place(x=120,y=280)
#================================================= fim da opção 2 ============================================================

#=================================================== OPÇÃO 3 =================================================================

    texto_g = tk.CTkLabel(opcao.tab('OPÇÃO 3'),text='% DE GORDURA\n''\n'
                                  ' BF é a sigla para “body fat” que em português pode\n'
                                  ' ser entendido como percentual de gordura corporal.\n'
                                  ' Em resumo esse percentual mede a proporção que existe\n'
                                  ' de gordura e músculo em nosso corpo !',font=('system',12))
    texto_g.place(x=20,y=10)

    sexo_g = tk.IntVar()
    peso_g = tk.IntVar()
    altura_g = tk.IntVar()
    idade_g = tk.IntVar()

   

    def calcular_composicao_corporal():
      sexo_var = sexo_g.get()
      peso_var = peso_g.get()
      altura_var = altura_g.get()
      idade_var = idade_g.get()
    
      if sexo_var == 1:
        bf = (1.20 * (float(peso_var) / float(altura_var)**2)) + (0.23 * float(idade_var)) - 16.2
        massa_gorda = float(bf) / 100 * float(peso_var)
        massa_magra = float(peso_var) - float(massa_gorda)
        
        exi_bf = tk.CTkLabel(opcao.tab('OPÇÃO 3'),text=f'Seu bf: {bf:.2f}',font=('system',12.5))
        exi_bf.place(x=90,y=350)
        result_bf = tk.CTkLabel(opcao.tab('RESULT'),text=f' {bf:.2f}',font=('impact',15))#EXIBIR RESULTADO NA TELA DE RESULTADO 
        result_bf.place(x=35,y=210)

        exi_magra = tk.CTkLabel(opcao.tab('OPÇÃO 3'),text=f'Sua massa magra: {massa_magra:.2f} %',font=('system',12.5))
        exi_magra.place(x=90,y=380)
        result_magra = tk.CTkLabel(opcao.tab('RESULT'),text=f' {massa_magra:.2f} %',font=('impact',15))#EXIBIR RESULTADO NA TELA DE RESULTADO 
        result_magra.place(x=105,y=160)

        exi_gorda = tk.CTkLabel(opcao.tab('OPÇÃO 3'),text=f'Sua massa gorda: {massa_gorda:.2f} %',font=('system',15))
        exi_gorda.place(x=90,y=410)
        result_gorda = tk.CTkLabel(opcao.tab('RESULT'),text=f' {massa_gorda:.2f} %',font=('impact',15))#EXIBIR RESULTADO NA TELA DE RESULTADO 
        result_gorda.place(x=105,y=110)
        
      elif sexo_var  == 2:
        bf = float((1.20 * (float(peso_var) / float(peso_var)**2)) + (0.23 * float(peso_var)) - 5.4)
        massa_gorda = float(bf) / 100 * float(peso_var)
        massa_magra = float(peso_var) - float(massa_gorda)
        
        exi_bf = tk.CTkLabel(opcao.tab('OPÇÃO 3'),text=f'Seu bf: {bf:.2f}',font=('system',12.5))
        exi_bf.place(x=90,y=350)
        result_bf = tk.CTkLabel(opcao.tab('RESULT'),text=f' {bf:.2f}',font=('impact',15))#EXIBIR RESULTADO NA TELA DE RESULTADO 
        result_bf.place(x=35,y=210)

        
        exi_magra = tk.CTkLabel(opcao.tab('OPÇÃO 3'),text=f'Sua massa magra: {massa_magra:.2f} %',font=('system',12.5))
        exi_magra.place(x=90,y=380)
        result_magra = tk.CTkLabel(opcao.tab('RESULT'),text=f' {massa_magra:.2f} %',font=('impact',15))#EXIBIR RESULTADO NA TELA DE RESULTADO 
        result_magra.place(x=105,y=160)

        exi_gorda = tk.CTkLabel(opcao.tab('OPÇÃO 3'),text=f'Sua massa gorda: {massa_gorda:.2f} %',font=('system',12.5))
        exi_gorda.place(x=90,y=410)
        result_gorda = tk.CTkLabel(opcao.tab('RESULT'),text=f' {massa_gorda:.2f} %',font=('impact',15))#EXIBIR RESULTADO NA TELA DE RESULTADO 
        result_gorda.place(x=105,y=110)

    idade_g = tk.CTkEntry(opcao.tab('OPÇÃO 3'),fg_color='white', placeholder_text='idade',font=('System',15),
                              width=250,border_color='light green', border_width=2,corner_radius=15,
                              height=35, text_color='black')
    idade_g.place(x=90, y=130)

    peso_g = tk.CTkEntry(opcao.tab('OPÇÃO 3'),fg_color='white', placeholder_text='Peso',font=('System',15),
                              width=250,border_color='light green', border_width=2,corner_radius=15,
                              height=35, text_color='black')
    peso_g.place(x=90,y=170)


    altura_g = tk.CTkEntry(opcao.tab('OPÇÃO 3'),fg_color='white', placeholder_text='Altura',font=('System',15),
                              width=250,border_color='light green', border_width=2,corner_radius=15,
                              height=35, text_color='black')
    altura_g.place(x=90,y=210)


    sexo_M_g = tk.CTkRadioButton(opcao.tab('OPÇÃO 3'), text='MASCULINO',fg_color='dark green',hover_color='light green', variable=sexo_g, value=1)
    sexo_F_g = tk.CTkRadioButton(opcao.tab('OPÇÃO 3'), text='FEMININO',fg_color='dark green',hover_color='light green',variable=sexo_g, value=2)
    sexo_M_g.place(x=110, y=260)
    sexo_F_g.place(x=225, y=260)

    botao_g = tk.CTkButton(opcao.tab('OPÇÃO 3'),text='>>START<<',hover_color='light green',fg_color='dark green',font=('System',15),border_width=2,
                           corner_radius=9,width=160,height=35,command=calcular_composicao_corporal)
    botao_g.place(x=120,y=300)
#=========================================================== FIM OPÇÃO 3 =============================================


#====================================================== OPÇÃO RESULT==================================================

    result = tk.CTkLabel(opcao.tab('RESULT'),text='RESULTADO',font=('IMPACT',25))
    result.place(x=170,y=10)

    resu_meta = tk.CTkLabel(opcao.tab('RESULT'),text='Metabolismo :',font=('impact',15))
    resu_meta.place(x=10,y=60)

    resu_gorda = tk.CTkLabel(opcao.tab('RESULT'),text='Massa gorda :',font=('impact',15))
    resu_gorda.place(x=10,y=110)

    resu_magra = tk.CTkLabel(opcao.tab('RESULT'),text='Massa magra :',font=('impact',15))
    resu_magra.place(x=10,y=160)

    resu_bf = tk.CTkLabel(opcao.tab('RESULT'),text='BF :',font=('impact',15))
    resu_bf.place(x=10,y=210)

    resu_imc = tk.CTkLabel(opcao.tab('RESULT'),text='IMC :',font=('impact',15))
    resu_imc.place(x=10,y=260)

    barra_divi = tk.CTkLabel(opcao.tab('RESULT'),text= 53*'-', text_color='light green', font=('system',35))
    barra_divi.place(x=5,y=310)
    
    dieta = tk.CTkLabel(opcao.tab('RESULT'),text='CRIA DIETA',font=('impact',26))
    dieta.place(x=170,y=330)
#=========================================================== JANELA DE DIETA ========================================

    

    def elinina_p(): # remove a janela_p
        janela_p.withdraw()
  
    def wrapper(): # função que execulta eliminar a tela (janela_p) e chama a tela de dieta 
          elinina_p()
          tela_dieta()
   
    def tela_dieta():
        janela_dieta = tk.CTk()
        janela_dieta.title('APP 0.1')
        janela_dieta.geometry('850x850+350+40')
        janela_dieta.maxsize(width=900,height=680) 
        janela_dieta.minsize(width=500,height=580)
   
       
#===================================== FUNÇÊS DE CRIA E LIMPA ==============================================
        ficha = pathlib.Path('DIETA.xlsx')  
         
        if ficha.exists():
              pass
        else:
           ficha=Workbook()
           folha=ficha.active
           # CAFÉ
           folha['A1']= 'Primeiro café'
           folha['B1']= 'Segunda café'
           folha['C1']= 'Terceiro café'
           folha['D1']= 'Lanche'
           folha['E1']= 'Lanche'
           folha['F1']= 'OBS do café'
          
           #Almoço
           folha['A11']= 'Primeiro almoço'
           folha['B11']= 'Segunda almoço'
           folha['C11']= 'Terceiro almoço'
           folha['D11']= 'Lanche'
           folha['E11']= 'Lanche'
           folha['F11']= 'OBS do almoço'
          
          #jantar
           folha['A20']= 'Primeiro jantar'
           folha['B20']= 'Segunda jantar'
           folha['C20']= 'Terceiro jantar'
           folha['D20']= 'Lanche'
           folha['E20']= 'Lanche'
           folha['F20']= 'OBS do jantar'
          
           ficha.save('DIETA.xlsx')   

        def enviar():
           
         # pegadno dados do cafe !
           p_cafe = cafe1.get()
           s_cafe = cafe2.get()
           t_cafe = cafe3.get()
           q_cafe = cafe4.get()
           c_cafe = cafe5.get()
           obs_cafe = cafe6.get(0.0, END)

         # pegando dados do almoço !
           p_almoço = almoço1.get()
           s_almoço = almoço2.get()
           t_almoço = almoço3.get()
           q_almoço = almoço4.get()
           c_almoço = almoço5.get()
           obs_almoço = almoço6.get(0.0, END)

         #  pegando dados do jantar !
           p_jantar = jantar0.get()
           s_jantar = jantar1.get()
           t_jantar = jantar3.get()
           q_jantar = jantar3.get()
           c_jantar = jantar4.get()
           obs_jantar = jantar6.get(0.0, END)
          
         # criando a planillha 
           ficha = openpyxl.load_workbook('DIETA.xlsx')
           folha = ficha.active
           #cafe posição
           folha.cell(column=1, row=2, value=p_cafe ) 
           folha.cell(column=2, row=2, value=s_cafe ) 
           folha.cell(column=3, row=2, value=t_cafe ) 
           folha.cell(column=4, row=2, value=q_cafe ) 
           folha.cell(column=5, row=2, value=c_cafe ) 
           folha.cell(column=6, row=2, value=obs_cafe ) 

           #almoço posição
           folha.cell(column=1, row=12, value=p_almoço ) 
           folha.cell(column=2, row=12, value=s_almoço ) 
           folha.cell(column=3, row=12, value=t_almoço ) 
           folha.cell(column=4, row=12, value=q_almoço ) 
           folha.cell(column=5, row=12, value=c_almoço ) 
           folha.cell(column=6, row=12, value=obs_almoço ) 

           #jantar posição 
           folha.cell(column=1, row=21, value=p_jantar ) 
           folha.cell(column=2, row=21, value=s_jantar ) 
           folha.cell(column=3, row=21, value=t_jantar ) 
           folha.cell(column=4, row=21, value=q_jantar ) 
           folha.cell(column=5, row=21, value=c_jantar ) 
           folha.cell(column=6, row=21, value=obs_jantar ) 

           ficha.save(r'DIETA.xlsx')
           messagebox.showinfo('SYSTEMA','Dados salvos com sucesso!')
           
         # Função limpa 
        def clear():
           #limpa cafe
           cafe1.delete(0,END)
           cafe2.delete(0,END)
           cafe3.delete(0,END)
           cafe4.delete(0,END)
           cafe5.delete(0,END)
           

           #limpa almoço
           almoço1.delete(0,END)
           almoço2.delete(0,END)
           almoço3.delete(0,END)
           almoço4.delete(0,END)
           almoço5.delete(0,END)

           #limpa jantar
           jantar0.delete(0,END)
           jantar1.delete(0,END)
           jantar3.delete(0,END)
           jantar4.delete(0,END)
           jantar5.delete(0,END)

          

        # Texto da variacel
        #cafe
        cafe1 = StringVar()
        cafe2 = StringVar()
        cafe3 = StringVar()
        cafe4 = StringVar()
        cafe5 = StringVar()
        
        #almoço
        almoço1 = StringVar()
        almoço2 = StringVar()
        almoço3 = StringVar()
        almoço4 = StringVar()
        almoço5 = StringVar()

        #jantar
        jantar0 = StringVar()
        jantar1 = StringVar()
        jantar3 = StringVar()
        jantar4 = StringVar()
        jantar5 = StringVar()
        
        def chmar_p():# função para chama a janela_p de volta e eliminar a janela_dieta.
            janela_p.deiconify()
            janela_dieta.withdraw()

#=========================================== CAFÉ / LANCHE ==========================================================
        lb_cafe = tk.CTkLabel(janela_dieta,text='CAFÉ/LANCHE',font=('IMPACT',15))
        lb_cafe.place(x=175,y=10)

        obs_cafe= tk.CTkLabel(janela_dieta,text='Observação:',font=('IMPACT',15))
        obs_cafe.place(x=630,y=10)
        
        #======================== 1 refeição ================================
        lb_cafe1 = tk.CTkLabel(janela_dieta,text='CAFÉ 1 :',font=('IMPACT',15))
        lb_cafe1.place(x=50,y=40)

        cafe1= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=cafe1, font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        cafe1.place(x=100,y=40)

        #======================== 2 refeição ==================================
        lb_cafe2 = tk.CTkLabel(janela_dieta,text='CAFÉ 2 :',font=('IMPACT',15))
        lb_cafe2.place(x=50,y=70)

        cafe2= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=cafe2,font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        cafe2.place(x=100,y=70)

        #======================== 3 refeição ================================== 
        lb_cafe3 = tk.CTkLabel(janela_dieta,text='CAFÉ 3 :',font=('IMPACT',15))
        lb_cafe3.place(x=50,y=100)

        cafe3= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=cafe3,font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        cafe3.place(x=100,y=100)

        #======================== 4 refeição ===================================
        lb_lanche = tk.CTkLabel(janela_dieta,text='LANCHE 1 :',font=('IMPACT',15))
        lb_lanche.place(x=35,y=130)

        cafe4= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=cafe4,font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        cafe4.place(x=100,y=130)

        #======================== 5 refeição ====================================
        lb_lanche2 = tk.CTkLabel(janela_dieta,text='LANCHE 2 :',font=('IMPACT',15))
        lb_lanche2.place(x=35,y=160)

        cafe5= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=cafe5,font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        cafe5.place(x=100,y=160)

        #======================= textbox ===========================================

        cafe6= tk.CTkTextbox(janela_dieta,fg_color='white',text_color='black',width=350,height=135
                             ,border_color='light green',border_width=3)
        cafe6.place(x=480,y=40)


#=========================================== ALMOÇO / LANCHE ==========================================================
        lb_almoço = tk.CTkLabel(janela_dieta,text='ALMOÇO/LANCHE',font=('IMPACT',15))
        lb_almoço.place(x=175,y=200)

        obs_almoço= tk.CTkLabel(janela_dieta,text='Observação:',font=('IMPACT',15))
        obs_almoço.place(x=630,y=200)

        #======================== 1 refeição ================================
        lb_almoço1 = tk.CTkLabel(janela_dieta,text='ALMOÇO 1 :',font=('IMPACT',15))
        lb_almoço1.place(x=30,y=230)

        almoço1= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=almoço1,font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        almoço1.place(x=100,y=230)

        #======================== 2 refeição ================================
        lb_almoço2 = tk.CTkLabel(janela_dieta,text='ALMOÇO 2 :',font=('IMPACT',15))
        lb_almoço2.place(x=30,y=260)

        almoço2= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=almoço2,font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        almoço2.place(x=100,y=260)

        #======================== 3 refeição ================================
        lb_almoço3 = tk.CTkLabel(janela_dieta,text='ALMOÇO 3 :',font=('IMPACT',15))
        lb_almoço3.place(x=30,y=290)

        almoço3= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=almoço3,font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        almoço3.place(x=100,y=290)

        #======================== 4 refeição ================================
        lb_almoço4= tk.CTkLabel(janela_dieta,text='LANCHE 1 :',font=('IMPACT',15))
        lb_almoço4.place(x=30,y=320)

        almoço4= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=almoço4,font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        almoço4.place(x=100,y=320)

        #======================== 5 refeição ================================
        lb_almoço5 = tk.CTkLabel(janela_dieta,text='LANCHE 2 :',font=('IMPACT',15))
        lb_almoço5.place(x=30,y=350)

        almoço5= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=almoço5,font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        almoço5.place(x=100,y=350)

        #======================= textbox ===========================================
        almoço6= tk.CTkTextbox(janela_dieta,fg_color='white',text_color='black',width=350,height=135
                             ,border_color='light green',border_width=3)
        almoço6.place(x=480,y=230)

#============================================== JANTAR / LANCHE =======================================================
        lb_jantar = tk.CTkLabel(janela_dieta,text='JANTAR/LANCHE',font=('IMPACT',15))
        lb_jantar.place(x=175,y=390)

        obs_jantar= tk.CTkLabel(janela_dieta,text='Observação:',font=('IMPACT',15))
        obs_jantar.place(x=630,y=390)

        #======================== 1 refeição ================================
        lb_jantar0 = tk.CTkLabel(janela_dieta,text='JANTAR 1 :',font=('IMPACT',15))
        lb_jantar0.place(x=30,y=420)
        jantar0= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=jantar0,font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        jantar0.place(x=100,y=420)

        #======================== 2 refeição ================================
        lb_jantar1 = tk.CTkLabel(janela_dieta,text='JANTAR 2 :',font=('IMPACT',15))
        lb_jantar1.place(x=30,y=450)
        jantar1= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=jantar1, font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        jantar1.place(x=100,y=450)

        #======================== 3 refeição ================================
        lb_jantar3 = tk.CTkLabel(janela_dieta,text='JANTAR 3:',font=('IMPACT',15))
        lb_jantar3.place(x=30,y=480)
        jantar3= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=jantar3, font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        jantar3.place(x=100,y=480)

        #======================== 4 refeição ================================
        lb_jantar4 = tk.CTkLabel(janela_dieta,text='LANCHE 1 :',font=('IMPACT',15))
        lb_jantar4.place(x=30,y=510)
        jantar4= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=jantar4,font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        jantar4.place(x=100,y=510)

        #======================== 5 refeição ================================
        lb_jantar5 = tk.CTkLabel(janela_dieta,text='LANCHE 2 :',font=('IMPACT',15))
        lb_jantar5.place(x=30,y=540)
        jantar5= tk.CTkEntry(janela_dieta,fg_color='white',textvariable=jantar5, font=('System',15),
                              width=350,height=25,border_color='light green', text_color='black')
        jantar5.place(x=100,y=540)

         #======================= textbox ===========================================

        jantar6= tk.CTkTextbox(janela_dieta,fg_color='white',text_color='black',width=350,height=135
                             ,border_color='light green',border_width=3)
        jantar6.place(x=480,y=420)

                                   #botão para joga a dieta no execel e limpa campos
        botao_gera = tk.CTkButton(janela_dieta,text='>>CRIAR<<',hover_color='light green',fg_color='dark green',font=('System',15),border_width=2,
                           corner_radius=9,width=160,height=35,command=enviar)
        botao_gera.place(x=350,y=590)

        botao_limpa = tk.CTkButton(janela_dieta,text='LIMPAR',hover_color='light green',fg_color='dark green',font=('System',15),border_width=2,
                           corner_radius=9,width=160,height=35,command=clear)
        botao_limpa.place(x=350,y=630)

        botao_voltar = tk.CTkButton(janela_dieta,text='< voltar',hover_color='#A6A6A6',fg_color='#545454',font=('System',15),border_width=2,
                           corner_radius=9,width=10,height=25,command=chmar_p)
        botao_voltar.place(x=15,y=10)
       

        janela_dieta.mainloop()

#========================================BOTÃO PARA ABRE A JANELA DA DIETA ============================================
    botao_dieta=tk.CTkButton(opcao.tab('RESULT'),text='CRIA DIETA',hover_color='light green',fg_color='dark green',font=('System',15),border_width=2,
                           corner_radius=9,width=160,height=35,command=wrapper)
    botao_dieta.place(x=143,y=380)   

    janela_p.mainloop()

   
        
    
#================================================== Area de login e registro =================================================
tabelas = tk.CTkTabview(janela,width=400, height=450,border_color='light green',border_width=2,
                                   corner_radius=20,segmented_button_unselected_color='green',
                                   segmented_button_fg_color='light green',segmented_button_unselected_hover_color='light green',
                                   segmented_button_selected_color='dark green',segmented_button_selected_hover_color='light green')
tabelas.add('Login')
tabelas.add('Registra')
tabelas.tab("Login")
tabelas.tab('Registra')
tabelas.place(x=225,y=35)

#----------------------Elemento do login--------------------------------------------------------------


def verifica_login():
    usuario = usuario_login.get()
    senha = senha_login.get()

    conn = sqlite3.connect('sistema_cadastro.db')
    cursor = conn.cursor()

    cursor.execute("""SELECT * FROM usuarios WHERE usuario = ? AND senha = ?""", (usuario, senha))

    verificar_dados = cursor.fetchone()

    try:
        if usuario == "" or senha == "":
            messagebox.showerror("Login", "Informe seus dados nos campos!")
        elif verificar_dados is not None:
            messagebox.showinfo("Login", "Login bem-sucedido!")
            conn.close()
            tela_p()
            janela.destroy()
        else:
            messagebox.showerror("Login", "Credenciais inválidas.")
            conn.close()

    except:
        messagebox.showerror("Login", "Erro ao verificar credenciais.")
        conn.close()
        

usuario_login = tk.CTkEntry(tabelas.tab('Login'),fg_color='white', placeholder_text='Usuario',font=('System',15),
                              width=250,border_color='light green', border_width=3.5,corner_radius=15,
                              height=35, text_color='black')
usuario_login.place(x=64,y=160)

senha_login = tk.CTkEntry(tabelas.tab('Login'),fg_color='white',placeholder_text='Senha',font=('Systeusm',15),
                             width=250,border_color='light green',border_width=3.5,corner_radius=15 ,
                             height=35, text_color='black',show='*')
senha_login.place(x=64,y=200)

lembra = tk.CTkSwitch(tabelas.tab('Login'),text='Lembra senha.',progress_color='dark green',font=('Arial bold',15))
lembra.place(x=64,y=245)

botao = tk.CTkButton(tabelas.tab('Login'),text='ENTRAR',width=160,height=35,fg_color='dark green',hover_color='light green',font=('system',12),
                                 border_width=2,corner_radius=15,command=verifica_login)
botao.place(x=109,y=290)

botao = tk.CTkButton(tabelas.tab('Login'),text='SAIR',width=160,height=35,fg_color='dark green',hover_color='light green',font=('system',12),
                                 border_width=2,corner_radius=15,command= janela.destroy)
botao.place(x=109,y=330)

#============================================ escohla de TEMA ==================================================

def chagen_appm(nova_aparencia):
  tk.set_appearance_mode(nova_aparencia)


temas=tk.CTkLabel(janela,text='TEMAS',bg_color='transparent',text_color=['#000','#fff'])
temas.place(x=50,y=430)
opcao_tema= tk.CTkOptionMenu(janela,values=['dark','light'],fg_color='green',button_color='dark green',button_hover_color='light green'
                             ,width=80,corner_radius=5 ,dropdown_fg_color='green',font=('impcat',15),command=chagen_appm)
opcao_tema.place(x=50,y=460)


#----------------------------imagem login-------------------------------

imagem = Image.open("login (1).png")
imagem = imagem.resize((165, 155), Image.ANTIALIAS)
imagem = ImageTk.PhotoImage(imagem)

imagem_label = tk.CTkLabel(tabelas.tab('Login'), image=imagem,text='')
imagem_label.place(x=85,y=5)

# ----------------------------------- ELEMENTO DE REGISTRO---------------------------------------------------

sexo_var = tk.StringVar()

def conecte_cadastra():
    # Conectando ao banco de dados
    conn = sqlite3.connect('sistema_cadastro.db')
    cursor = conn.cursor()

    # Criando tabela se ela não existir
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS usuarios(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario TEXT NOT NULL,
        email TEXT NOT NULL,
        senha TEXT NOT NULL,
        con_senha TEXT NOT NULL,
        sexo TEXT NOT NULL
    );
    """)

    conn.commit()

    usuario_cadas = usurio_r.get()
    email_cadas = email.get()  
    senha_cadas = senha.get()
    confirma_cadas = senha_c.get()
    sexo_coda = sexo_var.get()

    print(usuario_cadas)

    try:
        # Exibindo erros no cadastro
        if usuario_cadas == "" or email_cadas == "" or senha_cadas == "" or confirma_cadas == "" or sexo_coda == "":
            messagebox.showerror(title="SISTEMA DE REGISTRO", message="ERRO: Por favor, preencha todos os campos!")
        elif len(usuario_cadas) < 4:
            messagebox.showinfo(title="SISTEMA DE REGISTRO", message="O nome de usuário deve ter pelo menos 4 caracteres!")
        elif senha_cadas != confirma_cadas:
            messagebox.showerror(title="SISTEMA DE REGISTRO", message="ERRO: A confirmação de senha é diferente da senha!")
        else:
            # Inserindo os dados na tabela
            cursor.execute("""
            INSERT INTO usuarios(usuario, email, senha, con_senha, sexo)
            VALUES(?, ?, ?, ?, ?)""", (usuario_cadas, email_cadas, senha_cadas, confirma_cadas, sexo_coda))

            conn.commit()
            messagebox.showinfo(title="SISTEMA DE REGISTRO", message=f"PARABÉNS {usuario_cadas}, CADASTRO FEITO COM SUCESSO!")

    except Exception as e:
        messagebox.showerror(title="SISTEMA DE REGISTRO", message="ERRO: Ocorreu um problema durante o cadastro. Tente novamente!")
        print(f"Erro: {e}")

    conn.close()

# Limpando os campos após o cadastro!
def limpa_cadastro():
    usurio_r.delete(0, "end")
    email.delete(0, "end")
    senha.delete(0, "end")
    senha_c.delete(0, "end")

def cadastra_limpa_campos():
    conecte_cadastra()
    limpa_cadastro()
    

registra_lb=tk.CTkLabel(tabelas.tab('Registra'),text='REGISTRA',font=('impact',30),text_color='#74FF82')
registra_lb.place(x=140,y=5)

usurio_r = tk.CTkEntry(tabelas.tab('Registra'),fg_color='white',placeholder_text='Usuario...',font=('System',15),
                              width=250,border_color='light green', border_width=3,corner_radius=15,
                              height=35, text_color='black')
usurio_r.place(x=75,y=60)


email = tk.CTkEntry(tabelas.tab('Registra'),fg_color='white',placeholder_text='Email...',font=('System',15),
                            width=250,border_color='light green',border_width=3,corner_radius=15 ,
                            height=35, text_color='black')
email.place(x=75,y=100)

senha = tk.CTkEntry(tabelas.tab('Registra'),fg_color='white',placeholder_text=' Senha...',show='*',font=('System',15),
                               width=250,border_color='light green',border_width=3,corner_radius=15 ,
                               height=35, text_color='black')
senha.place(x=75,y=140)

senha_c = tk.CTkEntry(tabelas.tab('Registra'),fg_color='white',placeholder_text='Confirma senha...',show='*',font=('System',15)
                             ,width=250,border_color='light green',border_width=3,corner_radius=15 ,
                             height=35, text_color='black')
senha_c.place(x=75,y=180)

b_termos = tk.CTkFrame(tabelas.tab('Registra'),width=295,height=40,border_color='light green',border_width=2,
                                   corner_radius=10).place(x=65,y=272)
termos = tk.CTkCheckBox(tabelas.tab('Registra'),fg_color='dark green',hover_color='light green',text=" concorda com todos os termos e politicas!")
termos.place(x=79,y=280)

sexo_M_g = tk.CTkRadioButton(tabelas.tab('Registra'), text='MASCULINO',fg_color='dark green',hover_color='light green',variable=sexo_var,value="M").place(x=100,y=230)
sexo_imc_F = tk.CTkRadioButton(tabelas.tab('Registra'),text='FEMININO',hover_color='light green',fg_color='dark green',variable=sexo_var,value="F").place(x=210,y=230)
#---------------------------------boto de registra-----------------------------

botao_r = tk.CTkButton(tabelas.tab('Registra'),text='CADASTRE-SE',fg_color='dark green',hover_color='light green',
                                 border_width=2,corner_radius=9,font=('System',15),width=160,height=35,command=cadastra_limpa_campos)
botao_r.place(x=125,y=335)



janela.mainloop()
##================================================== fim da area de login e registro =================================================